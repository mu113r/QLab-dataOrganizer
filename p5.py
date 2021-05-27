import openpyxl as xl

input_path = ""
output_path = ""

def file_name(each_input, wb_out, current_row):
    list_of_words = each_input.split('/')
    name = (list_of_words[-1])[:-5]
    ws_out = wb_out.active
    ws_out.cell(row=current_row, column=1, value=name)
    return wb_out
    
def build_first_row(wb_in):
    ws_in = wb_in['Summary']
    wb_out = xl.Workbook()
    ws_out = wb_out.active
    ws_out.cell(row=1, column=1, value=ws_in.cell(row=1, column=1).value)
    ws_out.cell(row=1, column=2, value=ws_in.cell(row=1, column=3).value)
    ws_out.cell(row=1, column=3, value=ws_in.cell(row=1, column=4).value)
    ws_out.cell(row=1, column=4, value=ws_in.cell(row=1, column=7).value)
    ws_out.cell(row=1, column=5, value=ws_in.cell(row=1, column=9).value)
    return wb_out

def cbodies(wb_in, wb_out, current_row):
    ws_out = wb_out.active
    ws_in = wb_in['Summary']
    ws_out.cell(row=current_row, column=2, value=ws_in.cell(row=2, column=3).value)
    ws_out.cell(row=current_row, column=3, value=ws_in.cell(row=2, column=4).value)
    ws_out.cell(row=current_row, column=4, value=ws_in.cell(row=2, column=7).value)
    ws_out.cell(row=current_row, column=5, value=ws_in.cell(row=2, column=9).value)
    return wb_out
    
def fix_width(wb_out):
    ws_out = wb_out.active
    #col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','AA','AB','AC','AD','AE']
    #for col in col_list:
    ws_out.column_dimensions['A'].width = 25
    ws_out.column_dimensions['B'].width = 22
    ws_out.column_dimensions['C'].width = 20
    ws_out.column_dimensions['D'].width = 30
    ws_out.column_dimensions['E'].width = 52    
    return wb_out

def build_output(list_of_input_path):
    current_row = 2
    wb_in = xl.load_workbook(list_of_input_path[0])
    wb_out = build_first_row(wb_in)
    for each_input in list_of_input_path:
        wb_in = xl.load_workbook(each_input)
        wb_out = file_name(each_input, wb_out, current_row)
        wb_out = cbodies(wb_in, wb_out, current_row)
        current_row += 1
        wb_out = fix_width(wb_out)
    wb_out.save('output.xlsx')
        #sheet_obj = wb_in[list_of_sheets[0]]
        #cell_obj = sheet_obj.cell(row = 1, column = 2)
        #colA = sheet_obj['A']
        #print(colA[0].value)
    

# wb_obj = xl.load_workbook(input_path)
# sheet_obj = wb_obj.active
# cell_obj = sheet_obj.cell(row = 1, column = 1)

# print(cell_obj.value)

