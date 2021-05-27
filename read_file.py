import openpyxl as xl
import statistics

input_path = ""
output_path = ""

def file_name(each_input, wb_out, current_row):
    list_of_words = each_input.split('/')
    name = (list_of_words[-1])[:-5]
    ws_out = wb_out.active
    ws_out.cell(row=current_row, column=1, value=name)
    return wb_out
    
def build_first_row(wb_in):
    ws_in = wb_in['Cell Body Contours']
    wb_out = xl.Workbook()
    ws_out = wb_out.active
    ws_out.cell(row=1, column=2, value=ws_in.cell(row=1, column=1).value)
    ws_out.cell(row=1, column=3, value=ws_in.cell(row=1, column=2).value)
    ws_out.cell(row=1, column=4, value=ws_in.cell(row=1, column=3).value)
    ws_in = wb_in['Cell Bodies Summary']
    ws_out.cell(row=1, column=5, value=ws_in.cell(row=1, column=3).value)
    ws_out.cell(row=1, column=6, value=ws_in.cell(row=1, column=4).value)
    ws_out.cell(row=1, column=7, value='n of 3D')
    ws_in = wb_in['Neuron Summary']
    i = 8
    for j in range(2, 5):
        for x in range(14, 22):
            ws_out.cell(row=1, column=i, value=(ws_in.cell(row=j, column=1).value + " " + ws_in.cell(row=1, column=x).value))
            i += 1
    return wb_out

def cbodies(wb_in, wb_out, current_row):
    ws_in = wb_in['Cell Body Contours']
    colA = ws_in['B']
    colB = ws_in['C']
    colC = ws_in['D']
    listA = []
    listB = []
    listC = []
    for cel in range(1, len(colA)):
        if (colA[cel].value != None) and (type(colA[cel].value) != type("")):
            listA.append(float(colA[cel].value))
    for cel in range(1, len(colB)):
        if (colB[cel].value != None) and (type(colB[cel].value) != type("")):
            listB.append(float(colB[cel].value))
    for cel in range(1, len(colC)):
        if (colC[cel].value != None) and (type(colC[cel].value) != type("")):
            listC.append(float(colC[cel].value))
    bigA = sorted(listA, reverse=True)[0]
    bigB = sorted(listB, reverse=True)[0]
    bigC = sorted(listC, reverse=True)[0]
    ws_out = wb_out.active
    ws_out.cell(row=current_row, column=2, value=bigA)
    ws_out.cell(row=current_row, column=3, value=bigB)
    ws_out.cell(row=current_row, column=4, value=bigC)
    return wb_out
    
def threedcbodies(wb_in, wb_out, current_row):
    ws_in = wb_in['Cell Bodies Summary']
    ws_out = wb_out.active
    colC = ws_in['C']
    colD = ws_in['D']
    listC = []
    listD = []
    for cel in range(1, len(colC)):
        if (colC[cel].value != None) and (type(colC[cel].value) != type("")):
            listC.append(float(colC[cel].value))
    for cel in range(1, len(colD)):
        if (colD[cel].value != None) and (type(colD[cel].value) != type("")):
            listD.append(float(colD[cel].value))
    ws_out.column_dimensions['E'].number_format = '#,##0.00'
    ws_out.column_dimensions['F'].number_format = '#,##0.00'        
    ws_out.cell(row=current_row, column=5, value=statistics.mean(listC))
    ws_out.cell(row=current_row, column=6, value=statistics.mean(listD))
    ws_out.cell(row=current_row, column=7, value=len(listC))
    return wb_out
    
def nsummary(wb_in, wb_out, current_row):
    ws_out = wb_out.active
    ws_in = wb_in['Neuron Summary']
    i = 8
    for j in range(2, 5):
        for x in range(14, 22):
            ws_out.cell(row=current_row, column=i, value=ws_in.cell(row=j, column=x).value)
            i += 1
    return wb_out
    
def fix_width(wb_out):
    ws_out = wb_out.active
    col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','AA','AB','AC','AD','AE']
    for col in col_list:
        ws_out.column_dimensions[col].width = 20
    return wb_out

def build_output(list_of_input_path):
    current_row = 2
    wb_in = xl.load_workbook(list_of_input_path[0])
    wb_out = build_first_row(wb_in)
    for each_input in list_of_input_path:
        wb_in = xl.load_workbook(each_input)
        wb_out = file_name(each_input, wb_out, current_row)
        wb_out = cbodies(wb_in, wb_out, current_row)
        wb_out = threedcbodies(wb_in, wb_out, current_row)
        wb_out = nsummary(wb_in, wb_out, current_row)
        current_row += 1
        wb_out = fix_width(wb_out)
    wb_out.save('output.xlsx')
        #sheet_obj = wb_in[list_of_sheets[0]]
        #cell_obj = sheet_obj.cell(row = 1, column = 2)
        #colA = sheet_obj['A']
        #print(colA[0].value)
    

# wb_obj = xl.load_workbook(input_path)
# sheet_obj = wb_obj.active
# cell_obj = sheet_obj.cell(row = 1, column = 1)

# print(cell_obj.value)

